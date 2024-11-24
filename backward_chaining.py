import pandas as pd
import yaml
from sklearn.metrics import confusion_matrix, accuracy_score, precision_score, f1_score, precision_recall_curve, auc
import matplotlib.pyplot as plt
import seaborn as sns
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import time

# The above libaries are defined in the readme.md file, and enable data handling mechanisms and statistical output of evaluation statistics

class DataLoader:
    """
    Class to handle loading and preprocessing the dataset.
    """
    def __init__(self, dataset_path):
        """
        Initialize the DataLoader with the dataset path.
        
        Args:
            dataset_path (str): The path to the CSV file containing the dataset.
        """
        self.df = pd.read_csv(dataset_path)
        # Convert all values other than 0 in the 'num' column to 1
        self.df['num'] = self.df['num'].apply(lambda x: 1 if x != 0 else 0)
    
    def get_data(self):
        """
        Get the preprocessed dataset.
        
        Returns:
            pd.DataFrame: The preprocessed dataset.
        """
        return self.df

class RuleLoader:
    """
    Class to handle loading the rules from a YAML file.
    """
    def __init__(self, rules_path):
        """
        Initialize the RuleLoader with the rules file path.
        
        Args:
            rules_path (str): The path to the YAML file containing the rules.
        """
        with open(rules_path, 'r') as file:
            self.rules = yaml.safe_load(file)
    
    def get_rules(self):
        """
        Get the rules loaded from the YAML file.
        
        Returns:
            dict: The loaded rules.
        """
        return self.rules

class InferenceEngine:
    """
    Class to handle the inference process using backward chaining.
    """
    def __init__(self, rules):
        """
        Initialize the InferenceEngine with the rules.
        
        Args:
            rules (dict): The rules to be used for inference.
        """
        self.rules = rules
    
    def evaluate_premise(self, facts, premise):
        """
        Evaluate a premise based on the given facts.
        
        Args:
            facts (dict): The known facts.
            premise (str): The premise to be evaluated.
        
        Returns:
            bool: The result of the premise evaluation.
        """
        try:
            return eval(premise, {}, facts)
        except Exception as e:
            # print(f"Error evaluating premise '{premise}': {e}")
            return False
    
    def backward_chain(self, goal, facts, rule_sets):
        """
        Apply the backward chaining process to infer facts.
        
        Args:
            goal (str): The goal to achieve.
            facts (dict): The known facts.
            rule_sets (dict): The set of rules to be used.
        
        Returns:
            dict: The inferred facts.
        """
        inferred_facts = facts.copy()
        
        if goal in inferred_facts:
            return inferred_facts
        
        for set_name, rule_set in rule_sets.items():
            for rule in rule_set['rules']:
                conclusion = rule['conclusion']
                if conclusion.startswith(goal):
                    premise = rule['premise']
                    fact_name, fact_value = conclusion.split('=')
                    fact_name = fact_name.strip()
                    fact_value = int(fact_value.strip())
                    
                    if self.evaluate_premise(facts, premise):
                        inferred_facts[fact_name] = fact_value
                        inferred_facts.update(self.backward_chain(fact_name, inferred_facts, rule_sets))
        
        return inferred_facts

class Predictor:
    """
    Class to handle the prediction process using the given rules.
    """
    def __init__(self, data, rules):
        """
        Initialize the Predictor with the data and rules.
        
        Args:
            data (pd.DataFrame): The dataset.
            rules (dict): The rules for making predictions.

        This function then starts from a goal and uses backward chaining to  check if the rules  can be verfied on the fact, and enable if teh goal is tru or not.
        """
        self.data = data
        self.rules = rules
        self.inference_engine = InferenceEngine(rules)
    
    def predict(self):
        """
        Apply the rules to the dataset to generate predictions.
        
        Returns:
            pd.DataFrame: The dataset with predictions appended.
        """
        for set_name, rule_set in self.rules['rule_sets'].items():
            predictions = []
            for _, row in self.data.iterrows():
                facts = row.to_dict()
                final_facts = self.inference_engine.backward_chain('prediction', facts, {set_name: rule_set})
                
                prediction = final_facts.get('prediction', 0)
                predictions.append(prediction)
            
            self.data[f'prediction_{set_name}'] = predictions
        return self.data

class MetricsAndPlotter:
    """
    Class to handle metric calculations and plotting.
    """
    def __init__(self, data, output_dir):
        """
        Initialize the MetricsAndPlotter with the data and output directory.
        
        Args:
            data (pd.DataFrame): The dataset with predictions.
            output_dir (str): The directory to save output files.
        """
        self.data = data
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def calculate_metrics(self):
        """
        Calculate metrics and generate the Precision-Recall curve.
        
        Returns:
            str: The path to the saved PR curve image.
        """
        pr_fig, pr_ax = plt.subplots()
        for set_name in self.data.columns[self.data.columns.str.startswith('prediction_')]:
            y_true = self.data['num']
            y_pred = self.data[set_name]
            
            y_pred = y_pred.apply(lambda x: 1 if x != 0 else 0)
            
            cm = confusion_matrix(y_true, y_pred)
            print(f"Confusion Matrix for {set_name}:")
            print(cm)

            if cm.shape == (2, 2):
                tn, fp, fn, tp = cm.ravel()
            else:
                tn = fp = fn = tp = 0
                
            accuracy = accuracy_score(y_true, y_pred)
            precision = precision_score(y_true, y_pred, zero_division=0)
            f1 = f1_score(y_true, y_pred, zero_division=0)
            
            print(f"Metrics for {set_name}:")
            print(f"True Positives: {tp}")
            print(f"True Negatives: {tn}")
            print(f"False Positives: {fp}")
            print(f"False Negatives: {fn}")
            print(f"Accuracy: {accuracy:.2f}")
            print(f"Precision: {precision:.2f}")
            print(f"F1 Score: {f1:.2f}")
            print()

            precision, recall, _ = precision_recall_curve(y_true, y_pred)
            pr_auc = auc(recall, precision)
            pr_ax.plot(recall, precision, marker='.', label=f'{set_name} (AUC = {pr_auc:.2f})')
        
        pr_ax.set_xlabel('Recall')
        pr_ax.set_ylabel('Precision')
        pr_ax.set_title('Precision-Recall Curve')
        pr_ax.legend()
        pr_curve_path = os.path.join(self.output_dir, 'pr_curve.png')
        pr_fig.savefig(pr_curve_path)
        plt.close(pr_fig)
        
        return pr_curve_path
    
    def save_to_excel(self, pr_curve_path):
        """
        Save the dataset with predictions and the PR curve to an Excel file.
        
        Args:
            pr_curve_path (str): The path to the PR curve image.
        """
        excel_path = os.path.join(self.output_dir, 'predictions.xlsx')
        self.data.to_excel(excel_path, index=False)
        
        wb = load_workbook(excel_path)
        ws = wb.create_sheet('PR Curve')
        img = Image(pr_curve_path)
        ws.add_image(img, 'A1')
        wb.save(excel_path)

def main():
    """
    Main function to run the entire pipeline.
    """

    #record execution start time
    start_time = time.time()

    dataset_path = 'processed.cleveland.data'
    rules_path = 'rules.yaml'
    output_dir = 'output_bwd'
    
    data_loader = DataLoader(dataset_path)
    data = data_loader.get_data()
    
    rule_loader = RuleLoader(rules_path)
    rules = rule_loader.get_rules()
    
    predictor = Predictor(data, rules)
    predictions = predictor.predict()
    
    metrics_and_plotter = MetricsAndPlotter(predictions, output_dir)
    pr_curve_path = metrics_and_plotter.calculate_metrics()
    metrics_and_plotter.save_to_excel(pr_curve_path)

    #calculate and print execution time
    print ("--- %s execution time ---" % (time.time() - start_time))

if __name__ == "__main__":
    main()
